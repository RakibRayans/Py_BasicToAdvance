import openpyxl as xl
from openpyxl.chart import BarChart, Reference

wb = xl.load_workbook("transection.xlsx")
sheet = wb["Sheet1"]
cell = sheet.cell(1, 1)
#print(cell.value)
#print(sheet.max_row)  #totle row of the sheet
#print(sheet.max_column)  #totle column of the sheet

for row in range(2, sheet.max_row + 1):
    cell = sheet.cell(row, 3)
    #print(cell.value)   #show the 3rd coloumn value
    corrected_price = cell.value * 0.9
    #print(corrected_price)    #printing the corrected values
    c_cell = sheet.cell(row, 4)
    c_cell.value = corrected_price

values = Reference(sheet,
          min_row=2,
          max_row=sheet.max_row,
          min_col=4,
          max_col=10)

chart = BarChart()
chart.add_data(values)
sheet.add_chart(chart, "f2")

wb.save("transection5.xlsx")