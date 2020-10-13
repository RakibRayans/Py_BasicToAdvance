import openpyxl as xl
from openpyxl.chart import BarChart, Reference

def process_wb(filename):
    wb = xl.load_workbook(filename)
    sheet = wb["Sheet1"]

    for row in range(2, sheet.max_row + 1)
        cell = sheet.cell(row, 3)
        c_price = sheet.cell * 0.9
        c_price_cell = sheet.cell(row, 4)
        c_price_cell.value = c_price

    values = Reference(sheet,
              min_row=2,
              max_row=sheet.max_row,
              min_col=4,
              max_col=sheet,max_col)
    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart, "f2")

    wb.save(filename)
